"""
API del panel administrativo, respaldada por la base de datos PROPIA.

Reemplaza al viejo proxy contra Tiendanube. Mantiene los mismos paths y la misma
forma de respuesta (ver serializers.py) para que el frontend siga funcionando.
"""
from __future__ import annotations

import html
import io
import logging
import re
import secrets
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import List, Optional

import requests
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func
from sqlalchemy.orm import Session, selectinload

from .auth import get_current_admin
from core import storage
from core.config import settings
from core.db import get_db
from core.home_config import (
    DEFAULTS as HOME_DEFAULTS,
    get_home_config,
    reset_home_config,
    save_home_config,
)
from core.models import (
    Category, Order, Product, ProductImage, Reserva, Variant, Setting,
    RESERVA_STATUSES,
)
from core.categorizar import categorizar_producto, marca_canonica
from core.quitar_fondo import disponible as recorte_disponible
from core.quitar_fondo import quitar_fondo as quitar_fondo_ia
from core.sanitize import clean_description
from .serializers import order_to_tn, product_to_tn, reserva_to_dict

log = logging.getLogger("panel_api")

# Todo el panel exige sesión de admin.
router = APIRouter(prefix="/api", tags=["panel"], dependencies=[Depends(get_current_admin)])

# static de la TIENDA (web-tienda/static): ahí viven las fotos locales que se
# sirven como /static/products/... — el path viejo duplicaba "web-tienda" (de
# cuando el panel era un paquete hermano) y el guardado local daba 500.
STORE_STATIC = Path(__file__).resolve().parent.parent / "static"


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))



def _rotar_bytes(content: bytes, grados: int) -> bytes:
    """Rota una imagen en memoria. Si algo falla devuelve la original.

    `expand=True` para que no recorte al girar 90/270. El signo es negativo
    porque Pillow rota antihorario y en la interfaz el botón gira horario.
    """
    grados = grados % 360
    if not grados:
        return content
    try:
        from PIL import Image
        Image.MAX_IMAGE_PIXELS = 40_000_000     # anti bomba de descompresión
        src = Image.open(io.BytesIO(content))
        fmt = (src.format or "JPEG").upper()
        if fmt not in ("JPEG", "PNG", "WEBP"):
            fmt = "JPEG"
        out = src.rotate(-grados, expand=True)
        if fmt == "JPEG" and out.mode in ("RGBA", "P"):
            out = out.convert("RGB")
        buf = io.BytesIO()
        out.save(buf, format=fmt, quality=90)
        return buf.getvalue()
    except Exception:  # noqa: BLE001
        log.exception("No se pudo rotar la imagen; se guarda sin girar")
        return content


def _precio_valido(raw) -> Decimal:
    """Convierte a Decimal validando. Sin esto, "NaN", "Infinity", negativos y
    "1e999" entraban a la base y de ahí al monto que se le cobra al cliente."""
    try:
        d = Decimal(str(raw))
    except (InvalidOperation, TypeError, ValueError):
        raise HTTPException(400, "Precio inválido")
    if not d.is_finite() or d <= 0 or d > Decimal("100000000"):
        raise HTTPException(400, "Precio fuera de rango")
    return d.quantize(Decimal("0.01"))

def slugify(text: str, max_len: int = 80) -> str:
    t = _strip_accents(text or "").lower()
    t = re.sub(r"[^a-z0-9]+", "-", t).strip("-")
    return t[:max_len].rstrip("-") or "producto"


def unique_handle(db: Session, base: str) -> str:
    handle = base
    i = 1
    while db.query(Product).filter_by(handle=handle).first():
        i += 1
        handle = f"{base}-{i}"
    return handle


ALLOWED_IMAGE_EXT = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".avif"}
ALLOWED_IMAGE_MIME = {"image/jpeg", "image/png", "image/webp", "image/gif", "image/avif"}
MAX_IMAGE_BYTES = 8 * 1024 * 1024  # 8 MB
# Firmas (magic numbers) para confirmar que el archivo ES una imagen real.
_IMAGE_MAGIC = (b"\xff\xd8\xff", b"\x89PNG\r\n\x1a\n", b"GIF87a", b"GIF89a", b"RIFF")


def validate_image(filename: str, content_type: str, content: bytes) -> str:
    """Valida una imagen subida y devuelve un nombre de archivo SEGURO y aleatorio.
    Rechaza por tamaño, extensión, MIME y firma binaria. Renombrado aleatorio
    evita path traversal y colisiones."""
    if len(content) > MAX_IMAGE_BYTES:
        raise HTTPException(413, "Imagen demasiado grande (máx 8 MB)")
    if len(content) < 64:
        raise HTTPException(400, "Archivo vacío o corrupto")
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXT:
        raise HTTPException(415, f"Extensión no permitida: {ext or '(ninguna)'}")
    if content_type and content_type.lower() not in ALLOWED_IMAGE_MIME:
        raise HTTPException(415, f"Tipo no permitido: {content_type}")
    if not any(content.startswith(sig) for sig in _IMAGE_MAGIC):
        raise HTTPException(415, "El contenido no es una imagen válida")
    return f"{secrets.token_hex(16)}{ext if ext != '.jpeg' else '.jpg'}"


MAX_IMAGE_SIDE = 1600      # px: alcanza para zoom en detalle; una foto de celular trae 4000+
JPEG_WEB_QUALITY = 82


def comprimir_imagen(content: bytes) -> bytes:
    """Recomprime una foto a tamaño web ANTES de guardarla.

    Las fotos salen del celular en 4-5 MB / 4000px: servidas así, las cards de
    la tienda quedan negras o pintadas a medias en el teléfono. Acá se aplica
    la rotación EXIF, se limita el lado mayor a MAX_IMAGE_SIDE y se guarda en
    calidad web. GIF/AVIF (o cualquier cosa que Pillow no abra) vuelven tal
    cual: mejor foto pesada que upload roto.
    """
    try:
        from PIL import Image, ImageOps
    except ImportError:
        return content
    try:
        Image.MAX_IMAGE_PIXELS = 40_000_000  # anti bomba de descompresión
        img = Image.open(io.BytesIO(content))
        fmt = (img.format or "").upper()
        if fmt == "MPO":
            fmt = "JPEG"   # foto de iPhone: JPEG multi-frame, el primer frame es la foto
        if fmt not in ("JPEG", "PNG", "WEBP"):
            return content
        img = ImageOps.exif_transpose(img)
        if max(img.size) > MAX_IMAGE_SIDE:
            img.thumbnail((MAX_IMAGE_SIDE, MAX_IMAGE_SIDE), Image.LANCZOS)
        if fmt == "JPEG" and img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        buf = io.BytesIO()
        if fmt == "PNG":
            img.save(buf, format="PNG", optimize=True)
        else:
            img.save(buf, format=fmt, quality=JPEG_WEB_QUALITY, optimize=True)
        out = buf.getvalue()
        # Si no achicó (foto ya optimizada), se queda la original.
        return out if len(out) < len(content) else content
    except Exception:  # noqa: BLE001 — nunca romper la subida por la compresión
        log.exception("comprimir_imagen falló; se sube la original")
        return content


def store_image_bytes(content: bytes, content_type: str, handle: str, safe_name: str) -> tuple[str, str | None]:
    """Guarda la imagen en Supabase Storage (si está configurado) o en disco local.
    Devuelve (src, local_path). En Supabase, src = URL pública y local_path = None."""
    content = comprimir_imagen(content)
    rel = f"products/{handle}/{safe_name}"
    if storage.is_enabled():
        url = storage.upload_bytes(content, rel, content_type or "image/jpeg")
        return url, None
    dest = STORE_STATIC / rel
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(content)
    return f"/static/{rel}", f"/static/{rel}"


def _productos_completos(db: Session):
    """Query de Product con variantes, imágenes y categorías cargadas de una.

    El serializer toca esas tres relaciones por producto. Sin carga anticipada,
    /products/all (que devuelve TODO el catálogo) hacía ~3 consultas por
    producto: cientos de viajes a la Postgres remota. Con esto son 4 en total.
    """
    from sqlalchemy.orm import selectinload
    return (db.query(Product)
            .options(selectinload(Product.variants),
                     selectinload(Product.images),
                     selectinload(Product.categories)))


def current_usd_rate(db: Session) -> float:
    s = db.get(Setting, "usd_rate")
    if s and s.value and s.value.get("rate"):
        try:
            return float(s.value["rate"])
        except (TypeError, ValueError):
            pass
    return settings.USD_TO_ARS_RATE


# --------------------------------------------------------------------------- #
# Productos
# --------------------------------------------------------------------------- #
@router.get("/store")
def store_info(db: Session = Depends(get_db)):
    """Datos de la tienda para el frontend del panel.

    El panel viejo tenía la URL de Tiendanube hardcodeada; ahora sale de la
    config para que los links apunten a NUESTRA tienda.
    """
    base = settings.STORE_BASE_URL.rstrip("/")
    return {
        "name": "MIAMI IMPORT",
        "url": base,
        "product_url_base": f"{base}/productos/",
        "usd_rate": current_usd_rate(db),
    }


# Los productos "a pedido" no son stock real: se listan aparte, así que el
# catálogo normal del panel los excluye. `is_(False)` cubre también las filas
# viejas donde la columna quedó en NULL antes de existir el flag.
def _no_a_pedido(query):
    return query.filter((Product.a_pedido.is_(False)) | (Product.a_pedido.is_(None)))


@router.get("/products")
def list_products(q: Optional[str] = None, page: int = 1, per_page: int = 200,
                  db: Session = Depends(get_db)):
    query = _no_a_pedido(_productos_completos(db))
    if q:
        like = f"%{q.lower()}%"
        query = query.filter(func.lower(Product.name).like(like) | func.lower(Product.brand).like(like))
    items = query.order_by(Product.id.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return [product_to_tn(p) for p in items]


@router.get("/products/all")
def list_all_products(db: Session = Depends(get_db)):
    q = _no_a_pedido(_productos_completos(db)).order_by(Product.id.desc())
    return [product_to_tn(p) for p in q.all()]


@router.get("/products/a_pedido")
def list_a_pedido(db: Session = Depends(get_db)):
    q = (_productos_completos(db)
         .filter(Product.a_pedido.is_(True))
         .order_by(Product.id.desc()))
    return [product_to_tn(p) for p in q.all()]


@router.get("/products/{pid}")
def get_product(pid: int, db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    return product_to_tn(p)


@router.put("/products/{pid}")
def update_product(pid: int, body: dict, db: Session = Depends(get_db)):
    """Actualiza datos del producto (nombre, marca, descripción, publicado)."""
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    if body.get("name") is not None:
        nombre = str(body["name"]).strip()[:500]
        if not nombre:
            raise HTTPException(400, "El nombre no puede quedar vacío")
        p.name = nombre
    if "brand" in body:
        p.brand = str(body.get("brand") or "").strip()[:255] or None
    if "description" in body:
        # Se renderiza con `| safe` en la tienda: sanitizar siempre.
        p.description = clean_description(str(body.get("description") or ""))
    if "published" in body:
        p.published = bool(body["published"])
    if "destacado" in body:
        p.destacado = bool(body["destacado"])
    if "mas_vendido" in body:
        p.mas_vendido = bool(body["mas_vendido"])

    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.post("/products/{pid}/variants")
def add_variant(pid: int, body: dict, db: Session = Depends(get_db)):
    """Agrega un talle nuevo a un producto existente.

    Se crea con stock 0 por defecto: solo deja la opción disponible, sin tocar
    las variantes que el producto ya tiene cargadas.
    """
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    talle = str(body.get("talle") or body.get("value") or "").strip().upper()[:255]
    if not talle:
        raise HTTPException(400, "Falta el talle")
    if any((v.value or "").strip().upper() == talle for v in p.variants):
        raise HTTPException(409, f"El talle {talle} ya existe en este producto")

    try:
        stock = max(0, int(body.get("stock", 0) or 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Stock inválido")

    # Precio: el del body, o el de la primera variante existente.
    price = None
    if body.get("price") not in (None, ""):
        price = _precio_valido(body["price"])
    base = p.variants[0] if p.variants else None
    if price is None:
        price = base.price if base else Decimal("0")

    v = Variant(
        product_id=p.id, value=talle, stock=stock, price=price,
        usd_price=(base.usd_price if base else None),
        compare_at_price=(base.compare_at_price if base else None),
        position=(max((x.position or 0) for x in p.variants) + 1) if p.variants else 1,
        visible=True,
    )
    db.add(v)
    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.delete("/variants/{pid}/{vid}")
def delete_variant(pid: int, vid: int, db: Session = Depends(get_db)):
    """Quita un talle del producto.

    No se borra si es el único que queda (un producto sin variantes no se puede
    comprar ni mostrar precio) ni si está comprometido en un pedido pendiente
    de pago, porque ahí hay stock reservado que quedaría sin dueño.
    """
    from core.models import Order, OrderItem

    v = db.get(Variant, vid)
    if not v or v.product_id != pid:
        raise HTTPException(404, "Talle no encontrado")

    p = db.get(Product, pid)
    if p and len(p.variants) <= 1:
        raise HTTPException(409, "Es el único talle del producto: no se puede quitar")

    reservado = (db.query(OrderItem)
                 .join(Order, Order.id == OrderItem.order_id)
                 .filter(OrderItem.variant_id == vid,
                         Order.payment_status == "pending",
                         Order.stock_reserved.is_(True))
                 .first())
    if reservado:
        raise HTTPException(409, "Hay un pedido pendiente con este talle. "
                                 "Resolvelo antes de quitarlo.")

    db.delete(v)
    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.put("/products/{pid}/images/orden")
def reorder_images(pid: int, body: dict, db: Session = Depends(get_db)):
    """Reordena la galería del producto.

    `ids` viene con el orden nuevo, de la portada a la última. La primera es la
    que se ve en el listado de la tienda y como foto principal del producto.
    Se exige la lista COMPLETA para no dejar posiciones repetidas: la relación
    ordena por `position` y con empates el orden que ve el cliente es al azar.
    """
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")

    ids = body.get("ids")
    if not isinstance(ids, list) or not ids:
        raise HTTPException(400, "Falta el orden de las fotos")
    try:
        ids = [int(i) for i in ids]
    except (TypeError, ValueError):
        raise HTTPException(400, "Orden inválido")

    actuales = {im.id: im for im in p.images}
    if set(ids) != set(actuales) or len(ids) != len(actuales):
        raise HTTPException(400, "El orden no coincide con las fotos del producto")

    for pos, iid in enumerate(ids, 1):
        actuales[iid].position = pos
    db.commit()
    db.refresh(p)
    return product_to_tn(p)


@router.delete("/products/{pid}/images/{image_id}")
def delete_product_image(pid: int, image_id: int, db: Session = Depends(get_db)):
    """Borra una foto del producto.

    Se renumeran las posiciones de las que quedan (1..N) para no dejar huecos
    ni empates: con posiciones repetidas el orden en la tienda sale al azar.
    Si se borra la portada, la que quedó primera pasa a ser la principal.
    El archivo del bucket se elimina DESPUÉS del commit, y si falla no se
    aborta: peor caso queda un archivo huérfano, no una foto fantasma.
    """
    p = db.get(Product, pid)
    img = db.get(ProductImage, image_id)
    if not p or not img or img.product_id != pid:
        raise HTTPException(404, "Imagen no encontrada")

    old_path = storage.path_from_url(img.src)
    db.delete(img)
    db.flush()
    db.refresh(p)
    for pos, im in enumerate(sorted(p.images, key=lambda x: x.position or 0), 1):
        im.position = pos
    db.commit()

    if old_path:
        try:
            storage.delete_path(old_path)
        except Exception:  # noqa: BLE001
            pass
    db.refresh(p)
    return product_to_tn(p)


@router.post("/products/{pid}/images/{image_id}/rotate")
def rotate_product_image(pid: int, image_id: int, body: dict, db: Session = Depends(get_db)):
    """Rota una imagen ya subida y la reemplaza en el mismo lugar.

    A diferencia del panel viejo (que tenía que borrar y resubir a Tiendanube),
    acá se reescribe el archivo en nuestro bucket y se conserva la posición.
    """
    p = db.get(Product, pid)
    img = db.get(ProductImage, image_id)
    if not p or not img or img.product_id != pid:
        raise HTTPException(404, "Imagen no encontrada")
    try:
        degrees = int(body.get("degrees", 0) or 0) % 360
    except (TypeError, ValueError):
        degrees = 0
    if degrees == 0:
        return {"ok": True, "unchanged": True}
    if not storage.is_enabled():
        raise HTTPException(503, "Storage no configurado")

    try:
        from PIL import Image
    except ImportError:
        raise HTTPException(503, "Falta Pillow para rotar imágenes")

    # Se lee de `src`, NO de la property `url`: `url` devuelve local_path si
    # existe (las fotos migradas de TN tienen "/static/..." ahí, que no es
    # descargable) y además es de solo lectura, así que no se le puede asignar.
    origen = img.src or ""
    bucket = (settings.SUPABASE_URL or "").rstrip("/")
    if not (bucket and origen.startswith(bucket)):
        # Allow-list del origen: evita convertir esto en un SSRF si alguna vez
        # se permite cargar imágenes por URL arbitraria.
        raise HTTPException(400, "Solo se pueden rotar imágenes de nuestro bucket")

    try:
        resp = requests.get(origen, timeout=30, allow_redirects=False)
        resp.raise_for_status()
        if len(resp.content) > MAX_IMAGE_BYTES:
            raise HTTPException(413, "La imagen es demasiado grande")
        Image.MAX_IMAGE_PIXELS = 40_000_000  # anti bomba de descompresión
        src = Image.open(io.BytesIO(resp.content))
        fmt = (src.format or "JPEG").upper()
        if fmt not in ("JPEG", "PNG", "WEBP"):
            fmt = "JPEG"
        # expand=True para que no recorte al rotar 90/270.
        out = src.rotate(-degrees, expand=True)
        if fmt == "JPEG" and out.mode in ("RGBA", "P"):
            out = out.convert("RGB")
        buf = io.BytesIO()
        out.save(buf, format=fmt, quality=90)
        data = comprimir_imagen(buf.getvalue())
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(502, f"No se pudo procesar la imagen: {exc}")

    ext = {"JPEG": "jpg", "PNG": "png", "WEBP": "webp"}[fmt]
    rel = f"products/{slugify(p.name or 'producto')}/{secrets.token_hex(16)}.{ext}"
    try:
        new_url = storage.upload_bytes(
            data, rel, f"image/{'jpeg' if ext == 'jpg' else ext}")
    except RuntimeError as exc:
        raise HTTPException(502, f"No se pudo guardar la imagen rotada: {exc}")

    old_path = storage.path_from_url(img.src)
    img.src = new_url          # `url` es property de solo lectura
    img.local_path = None      # la copia local vieja ya no aplica
    db.commit()
    if old_path:
        try:
            storage.delete_path(old_path)  # recién ahora, con la nueva ya guardada
        except Exception:  # noqa: BLE001
            pass
    return {"ok": True, "src": new_url}


@router.delete("/products/{pid}")
def delete_product(pid: int, db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    # Borrar también las imágenes del bucket de Supabase (si aplica).
    if storage.is_enabled():
        for img in p.images:
            path = storage.path_from_url(img.src)
            if path:
                storage.delete_path(path)
    db.delete(p)
    db.commit()
    return {"ok": True}


@router.put("/variants/{pid}/{vid}/stock")
def update_variant_stock(pid: int, vid: int, body: dict, db: Session = Depends(get_db)):
    v = db.get(Variant, vid)
    if not v or v.product_id != pid:
        raise HTTPException(404, "Variante no encontrada")
    try:
        nuevo = int(body.get("stock", 0))
    except (TypeError, ValueError):
        raise HTTPException(400, "Stock inválido")
    if nuevo < 0:
        raise HTTPException(400, "El stock no puede ser negativo")
    v.stock = nuevo
    db.commit()
    return {"ok": True, "stock": v.stock}


@router.put("/variants/{pid}/{vid}")
def update_variant(pid: int, vid: int, body: dict, db: Session = Depends(get_db)):
    # `usd_price` y `price` se mantienen SIEMPRE coherentes entre sí usando la
    # cotización del panel: la tienda muestra el USD guardado como precio
    # principal, así que si quedaran desacoplados (editás pesos y el USD viejo
    # no se toca) el cliente ve un precio y paga otro. Editar cualquiera de los
    # dos recalcula el otro.
    v = db.get(Variant, vid)
    if not v or v.product_id != pid:
        raise HTTPException(404, "Variante no encontrada")
    try:
        rate = Decimal(str(current_usd_rate(db)))
        if "stock" in body:
            stock = int(body["stock"])
            if stock < 0:
                raise HTTPException(400, "El stock no puede ser negativo")
            v.stock = stock
        if "usd_price" in body and body["usd_price"] not in (None, ""):
            v.usd_price = _precio_valido(body["usd_price"])
            if rate > 0:
                v.price = (v.usd_price * rate).quantize(Decimal("0.01"))
        elif "price" in body:
            v.price = _precio_valido(body["price"])
            if v.usd_price is not None and rate > 0:
                v.usd_price = (v.price / rate).quantize(Decimal("0.01"))
        if "sku" in body:
            v.sku = (str(body["sku"]) or "")[:120] or None
        if "promotional_price" in body and body["promotional_price"] not in (None, ""):
            v.promotional_price = _precio_valido(body["promotional_price"])
    except (TypeError, ValueError):
        raise HTTPException(400, "Valor inválido")
    db.commit()
    return {"ok": True}


@router.post("/products")
async def create_product(
    name: str = Form(...),
    brand: str = Form(...),
    description: str = Form(""),
    # Precio opcional: en los productos "a pedido" el precio puede no saberse
    # todavía. En el catálogo normal sigue siendo obligatorio (se valida abajo).
    price: str = Form(""),
    talles: str = Form(""),
    stock_por_talle: int = Form(1),
    publicado: bool = Form(True),
    # Producto "a pedido": no es stock real ni se publica en la tienda. Queda en
    # la pestaña aparte del panel para tomar reservas.
    a_pedido: bool = Form(False),
    images: List[UploadFile] = File([]),
    convertir_a_ars: bool = Form(False),
    # Grados de giro por imagen, en el mismo orden que `images` ("0,90,0").
    # El formulario ya lo mandaba, pero el backend no lo recibía: la foto que
    # se veía derecha en la vista previa se subía igual de dada vuelta.
    rotations: str = Form(""),
    db: Session = Depends(get_db),
):
    talles_list = [t.strip().upper() for t in re.split(r"[,;]", talles) if t.strip()]
    base_handle = slugify(f"{brand}-{name}")
    handle = unique_handle(db, base_handle)
    sku_base = re.sub(r"[^A-Z0-9]", "", _strip_accents(f"{brand}{name}").upper())[:30] or "PROD"

    # Un producto "a pedido" nunca se publica en la tienda (no es stock real) y
    # su stock arranca en 0.
    if a_pedido:
        publicado = False
        stock_por_talle = 0

    rate = Decimal(str(current_usd_rate(db)))
    if not price.strip():
        if not a_pedido:
            raise HTTPException(400, "Falta el precio")
        precio_ars = None    # a pedido sin precio definido todavía
        usd_val = None
    else:
        precio_num = _precio_valido(price)
        if convertir_a_ars:
            usd_val = precio_num
            precio_ars = (precio_num * rate).quantize(Decimal("0.01"))
        else:
            precio_ars = precio_num.quantize(Decimal("0.01"))
            usd_val = (precio_ars / rate).quantize(Decimal("0.01")) if rate else None

    # La descripción se renderiza con `| safe` en la tienda, así que el HTML
    # que se guarda tiene que ser SOLO el que armamos acá: los valores del
    # formulario van escapados. Sin esto, un `<script>` guardado desde el panel
    # queda persistente en el dominio donde vive el checkout de Stripe.
    desc = f"<p>{html.escape(name)}</p>"
    if brand:
        desc += f"<p>Marca: {html.escape(brand)}</p>"
    desc += (f"<p>{html.escape(description)}</p>" if description
             else "<p>Producto original importado.</p>")

    prod = Product(name=name, handle=handle, description=desc, brand=brand,
                   published=publicado, a_pedido=a_pedido)
    db.add(prod)
    db.flush()

    # Categorización automática: sin esto el producto no aparece en el menú de
    # la tienda (pasó con 113 de 248 productos, el 45% del catálogo).
    if not a_pedido:
        try:
            categorizar_producto(db, prod)
            db.flush()
        except Exception:  # noqa: BLE001 — nunca frenar un alta por esto
            log.exception("no se pudo categorizar el producto %s", prod.id)

    talles_iter = talles_list or [None]
    for i, t in enumerate(talles_iter, 1):
        db.add(Variant(
            product_id=prod.id, price=precio_ars, usd_price=usd_val, stock=stock_por_talle,
            sku=f"{sku_base}-{slugify(t).upper()}" if t else sku_base, value=t, position=i,
        ))
    db.flush()

    imgs_ok = 0
    giros = [int(g) % 360 if g.strip().lstrip("-").isdigit() else 0
             for g in (rotations or "").split(",")]
    for i, upfile in enumerate(images, 1):
        content = await upfile.read()
        safe_name = validate_image(upfile.filename or "", upfile.content_type or "", content)
        # Aplicar el giro que el usuario le dio en la vista previa ANTES de
        # guardar: si no, la foto se sube tal como salió de la cámara.
        giro = giros[i - 1] if i - 1 < len(giros) else 0
        if giro:
            content = _rotar_bytes(content, giro)
        src, local_path = store_image_bytes(content, upfile.content_type or "", handle, safe_name)
        db.add(ProductImage(product_id=prod.id, src=src, local_path=local_path,
                            position=i, alt=name))
        imgs_ok += 1
    db.commit()

    return {
        "ok": True,
        "product_id": prod.id,
        "a_pedido": a_pedido,
        # Los "a pedido" no viven en la tienda: no tiene sentido un link público.
        "url": None if a_pedido else f"{settings.STORE_BASE_URL}/productos/{handle}/",
        "imagenes_subidas": imgs_ok,
        "variantes_creadas": len(talles_iter),
    }


@router.post("/products/{pid}/images")
async def upload_image(pid: int, file: UploadFile = File(...),
                       position: Optional[int] = Form(None),
                       db: Session = Depends(get_db)):
    p = db.get(Product, pid)
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    content = await file.read()
    safe_name = validate_image(file.filename or "", file.content_type or "", content)
    src, local_path = store_image_bytes(content, file.content_type or "", p.handle, safe_name)
    # Sin posición explícita, la foto va al final de la galería. Antes entraban
    # todas con 99: al subir varias de una quedaban empatadas y el orden que
    # veía el cliente en la tienda salía al azar.
    if position is None:
        position = (max((im.position or 0) for im in p.images) + 1) if p.images else 1
    db.add(ProductImage(product_id=p.id, src=src, local_path=local_path,
                        position=position, alt=p.name))
    db.commit()
    return {"ok": True, "src": src}


# --------------------------------------------------------------------------- #
# Pedidos / estadísticas
# --------------------------------------------------------------------------- #
@router.get("/orders/{oid}")
def get_order(oid: int, db: Session = Depends(get_db)):
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    return order_to_tn(o)


@router.post("/orders/{oid}/status")
def set_order_status(oid: int, body: dict, db: Session = Depends(get_db)):
    from core.models import ORDER_STATUSES
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    new = body.get("status")
    if new not in ORDER_STATUSES:
        raise HTTPException(400, f"Estado inválido. Opciones: {', '.join(ORDER_STATUSES)}")
    # "refunded" solo puede salir del endpoint de reembolso, que sí le pide la
    # plata a Stripe. Marcarlo a mano dejaba al cliente sin su devolución y con
    # el pedido figurando como reembolsado (contracargo asegurado a los 60 días).
    if new == "refunded":
        raise HTTPException(409, "Usá el botón de reembolso: marcarlo a mano no "
                                 "devuelve el dinero en Stripe")

    from checkout import _release_stock, _try_reserve
    anterior = o.status
    aviso = None

    # --- Cancelar devuelve la mercaderia a la venta -------------------------
    # El stock se descuenta al arrancar el checkout, no al cobrarse. Si al
    # cancelar no se devolvia, la prenda quedaba descontada PARA SIEMPRE:
    # figuraba agotada en la web sin haberse vendido nunca. El reembolso ya lo
    # hacia bien; el desplegable de estado, no. Diego lo vio y lo reporto.
    if new == "cancelled" and anterior != "cancelled":
        # Un pedido COBRADO no se cancela con el desplegable: devolver la
        # mercaderia y quedarse con la plata termina en un contracargo a los
        # ~60 dias, con multa. Para eso esta el boton de Reembolsar, que le
        # pide la devolucion a Stripe Y repone el stock.
        if o.payment_status == "paid":
            raise HTTPException(409, "Este pedido está cobrado. Usá el botón "
                                     "Reembolsar: devuelve la plata por Stripe "
                                     "y repone el stock.")
        devueltos = sum(it.quantity for it in o.items) if o.stock_reserved else 0
        _release_stock(db, o)          # idempotente: mira el flag stock_reserved
        # Un pedido cancelado no puede seguir figurando "esperando pago": deja
        # de aparecer en la lista de los que hay que ir a cobrar.
        if o.payment_status == "pending":
            o.payment_status = "cancelled"
        if devueltos:
            aviso = f"Volvieron {devueltos} unidad(es) al stock."

    # --- Reactivar un pedido cancelado tiene que volver a TOMAR la mercaderia
    # Si no, se promete algo que en el medio se le pudo haber vendido a otro.
    # Es todo o nada: si falta una unidad, no se reactiva y se dice cual falta.
    # 'backorder' = cobrado sin stock. Al pasarlo a Preparando hay que
    # DESCONTAR la prenda que se repuso; si no, esa unidad se revende.
    if anterior in ("cancelled", "backorder") and new not in ("cancelled", "backorder"):
        faltantes = _try_reserve(db, o)
        if faltantes:
            db.rollback()
            raise HTTPException(409, "No se puede reactivar: ya no queda stock de "
                                     + ", ".join(faltantes))
        # El pago vuelve a estar pendiente, no "cancelado": si no, el pedido
        # reactivado desaparecia de la lista de los que hay que ir a cobrar.
        if o.payment_status == "cancelled":
            o.payment_status = "pending"
        aviso = "Se volvió a descontar la mercadería del stock."

    o.status = new
    db.commit()
    return {"ok": True, "status": o.status, "payment_status": o.payment_status,
            "aviso": aviso}


@router.post("/orders/reconciliar")
def reconciliar_pagos(db: Session = Depends(get_db)):
    """Revisa los pedidos pendientes contra Stripe y acredita los ya cobrados.

    Existe porque el webhook puede no llegar (mal configurado, caído o
    demorado) y en ese caso el pedido queda "pendiente" con la plata cobrada.
    Acá se pregunta el estado real de cada PaymentIntent y se corrige.

    Es seguro de correr las veces que haga falta: solo toca pedidos que Stripe
    confirma como pagados y cuyo monto y moneda coinciden con lo esperado.
    """
    if not settings.STRIPE_SECRET_KEY:
        raise HTTPException(503, "Stripe no está configurado")
    try:
        import stripe
        stripe.api_key = settings.STRIPE_SECRET_KEY
    except Exception as exc:  # noqa: BLE001
        log.exception("reconciliar: no se pudo inicializar Stripe")
        raise HTTPException(502, f"No se pudo hablar con Stripe: {exc}") from exc

    from core.models import AuditLog, CartItem, Payment
    from checkout import _sd

    pendientes = (db.query(Order)
                  .filter(Order.payment_status == "pending")
                  .order_by(Order.id.desc()).limit(200).all())

    acreditados, sin_pagar, revisar, errores = [], [], [], []
    minor = settings.currency_minor_units

    for o in pendientes:
      try:
        pago = (db.query(Payment)
                .filter(Payment.order_id == o.id,
                        Payment.stripe_payment_intent_id.isnot(None))
                .order_by(Payment.id.desc()).first())
        if not pago:
            continue
        try:
            intent = _sd(stripe.PaymentIntent.retrieve(pago.stripe_payment_intent_id))
        except Exception as e:  # noqa: BLE001
            errores.append({"pedido": o.number, "error": str(e)[:120]})
            continue

        if intent.get("status") != "succeeded":
            # Guardar POR QUE no entro: sin esto el panel solo podia decir
            # "no pago", que es justo lo que a Diego no le alcanzaba.
            pago.estado_stripe = (intent.get("status") or "")[:40] or None
            err = _sd(intent.get("last_payment_error"))
            if err:
                pago.error_message = (err.get("message") or "")[:500] or None
                pago.error_code = ((err.get("decline_code") or err.get("code")
                                    or "")[:60]) or None
            sin_pagar.append({"pedido": o.number, "estado_stripe": intent.get("status")})
            continue

        cobrado = Decimal(intent.get("amount_received") or intent.get("amount") or 0) / minor
        moneda = (intent.get("currency") or "").upper()[:3]
        if cobrado != pago.amount or moneda != (pago.currency or "").upper()[:3]:
            pago.status = "review"
            o.payment_status = "review"
            pago.estado_stripe = (intent.get("status") or "")[:40] or None
            pago.error_message = (f"Cobrado {cobrado} {moneda}, "
                                  f"esperado {pago.amount} {pago.currency}")
            db.add(AuditLog(user_id=o.user_id, action="payment_amount_mismatch",
                            entity="order", entity_id=str(o.id)))
            revisar.append({"pedido": o.number, "cobrado": f"{cobrado} {moneda}",
                            "esperado": f"{pago.amount} {pago.currency}"})
            continue

        o.payment_status = "paid"
        o.status = "processing"
        pago.status = "paid"
        pago.raw = {"confirmado_por": "reconciliacion_manual"}
        # Mismo rastro que en el webhook: id del cobro, recibo de Stripe y
        # tarjeta. Un pedido acreditado a mano tiene que quedar tan probado
        # como uno acreditado solo.
        from checkout import _sellar_cobro
        _sellar_cobro(pago, intent)
        if o.cart_id:
            db.query(CartItem).filter(CartItem.cart_id == o.cart_id).delete(
                synchronize_session=False)
        db.add(AuditLog(user_id=o.user_id, action="payment_reconciled",
                        entity="order", entity_id=str(o.id)))
        acreditados.append({"pedido": o.number, "monto": f"{cobrado} {moneda}"})
      except Exception as exc:  # noqa: BLE001 — un pedido roto no frena el resto
        log.exception("reconciliar: falló el pedido %s", o.number)
        errores.append({"pedido": o.number, "error": f"{type(exc).__name__}: {str(exc)[:160]}"})

    # --- Por que NO pagaron (tambien los ya cancelados) ----------------------
    # El barrido pasa los abandonados a 'cancelled', asi que el bucle de arriba
    # (que solo mira 'pending') ya no los alcanza y nunca les escribia el
    # motivo. Diego abria un pedido cancelado y no tenia forma de saber si el
    # cliente habia intentado pagar — justo el dato que necesita para decidir
    # si sale a rescatar la venta.
    for o in (db.query(Order)
              .options(selectinload(Order.payments))
              .filter(Order.payment_status.in_(("cancelled", "failed")))
              .order_by(Order.id.desc()).limit(100).all()):
        pago = next((x for x in reversed(list(o.payments or []))
                     if x.stripe_payment_intent_id and not x.estado_stripe), None)
        if not pago:
            continue
        try:
            intent = _sd(stripe.PaymentIntent.retrieve(pago.stripe_payment_intent_id))
            pago.estado_stripe = (intent.get("status") or "")[:40] or None
            err = _sd(intent.get("last_payment_error"))
            if err:
                pago.error_message = (err.get("message") or "")[:500] or None
                pago.error_code = ((err.get("decline_code") or err.get("code")
                                    or "")[:60]) or None
        except Exception as exc:  # noqa: BLE001
            log.exception("reconciliar: no se pudo leer el motivo del pedido %s", o.number)
            errores.append({"pedido": o.number,
                            "error": f"motivo: {type(exc).__name__}: {str(exc)[:110]}"})

    # --- Sellado retroactivo -------------------------------------------------
    # Los pedidos que se cobraron ANTES de que guardaramos el rastro figuran
    # como pagados pero sin con que probarlo. Se les completa el id del cobro,
    # el recibo de Stripe y la tarjeta. NO se toca ningun estado: si ya estaba
    # pagado, sigue pagado; esto solo agrega la evidencia que falta.
    from checkout import _sellar_cobro
    sellados = []
    viejos = (db.query(Order)
              .options(selectinload(Order.payments))
              .filter(Order.payment_status == "paid")
              .order_by(Order.id.desc()).limit(200).all())
    for o in viejos:
        pago = next((x for x in reversed(list(o.payments or []))
                     # Se re-sella si falta CUALQUIERA de los datos: cada vez
                     # que agregamos un campo nuevo (medio, comision, neto) los
                     # cobros viejos quedan a medias, y una condicion que solo
                     # mire el charge_id no los vuelve a tocar nunca.
                     if x.stripe_payment_intent_id and not (
                         x.stripe_charge_id and x.metodo and x.neto is not None)), None)
        if not pago:
            continue
        try:
            intent = _sd(stripe.PaymentIntent.retrieve(pago.stripe_payment_intent_id))
            _sellar_cobro(pago, intent)
            if pago.stripe_charge_id:
                sellados.append({"pedido": o.number, "cobro": pago.stripe_charge_id})
        except Exception as exc:  # noqa: BLE001
            log.exception("reconciliar: no se pudo sellar el pedido %s", o.number)
            errores.append({"pedido": o.number,
                            "error": f"sellado: {type(exc).__name__}: {str(exc)[:120]}"})

    try:
        db.commit()
    except Exception as exc:  # noqa: BLE001
        db.rollback()
        log.exception("reconciliar: no se pudo guardar")
        raise HTTPException(500, f"No se pudieron guardar los cambios: {exc}") from exc
    return {
        "ok": True,
        "revisados": len(pendientes),
        "acreditados": acreditados,
        "sin_pagar": sin_pagar,
        "para_revisar": revisar,
        "sellados": sellados,
        "errores": errores,
    }



# --------------------------------------------------------------------------- #
# EL DINERO - lo que Diego le reclama a la LLC
# --------------------------------------------------------------------------- #
# La cuenta de Stripe esta a nombre de la LLC que se la abrio: la plata cae en
# el banco de ELLOS y despues se la giran a Diego. Para reclamar hay que poder
# decir, con numeros y con pruebas: cuanto entro, cuanto se llevo Stripe,
# cuanto quedo neto, y que Stripe YA lo deposito tal dia. Eso arma esta vista.

def _cobro_a_dict(o, pago) -> dict:
    medio = None
    if pago.card_brand and pago.card_last4:
        medio = f"{pago.card_brand.upper()} ....{pago.card_last4}"
    elif pago.metodo:
        from panel.serializers import _MEDIOS
        medio = _MEDIOS.get(pago.metodo, pago.metodo.replace("_", " ").title())
    return {
        "pedido": o.number,
        "order_id": o.id,
        "fecha": pago.paid_at.isoformat() if pago.paid_at else None,
        "cliente": o.contact_name,
        "email": o.email,
        "bruto": float(pago.amount or 0),
        "moneda": pago.currency,
        "comision": float(pago.fee) if pago.fee is not None else None,
        "neto": float(pago.neto) if pago.neto is not None else None,
        "moneda_neto": pago.moneda_liquidacion,
        "disponible_el": pago.disponible_el.isoformat() if pago.disponible_el else None,
        "medio": medio,
        "cobro_id": pago.stripe_charge_id,
        "recibo_url": pago.receipt_url,
    }


def _rango(desde: Optional[str], hasta: Optional[str]):
    """Convierte los filtros de fecha del panel. Vacio = todo."""
    d1 = d2 = None
    try:
        if desde:
            d1 = datetime.fromisoformat(desde).replace(tzinfo=timezone.utc)
        if hasta:
            d2 = datetime.fromisoformat(hasta).replace(
                tzinfo=timezone.utc, hour=23, minute=59, second=59)
    except ValueError as exc:
        raise HTTPException(400, "Fecha invalida: usa AAAA-MM-DD") from exc
    return d1, d2


@router.get("/dinero")
def dinero(desde: Optional[str] = None, hasta: Optional[str] = None,
           db: Session = Depends(get_db)):
    from core.models import Payment
    d1, d2 = _rango(desde, hasta)

    q = (db.query(Order, Payment)
         .join(Payment, Payment.order_id == Order.id)
         .filter(Payment.status == "paid")
         .order_by(Payment.paid_at.desc(), Order.id.desc()))
    if d1:
        q = q.filter(Payment.paid_at >= d1)
    if d2:
        q = q.filter(Payment.paid_at <= d2)

    cobros, bruto, comision, neto = [], 0.0, 0.0, 0.0
    monedas, monedas_neto, sin_datos = set(), set(), 0
    for o, pago in q.limit(500).all():
        cobros.append(_cobro_a_dict(o, pago))
        bruto += float(pago.amount or 0)
        if pago.currency:
            monedas.add(pago.currency.upper())
        if pago.neto is None:
            sin_datos += 1
        else:
            comision += float(pago.fee or 0)
            neto += float(pago.neto)
            if pago.moneda_liquidacion:
                monedas_neto.add(pago.moneda_liquidacion.upper())

    def _una(ms):
        if not ms:
            return None
        return ms.pop() if len(ms) == 1 else "mixta"

    resumen = {
        "cantidad": len(cobros),
        "bruto": round(bruto, 2),
        "moneda": _una(monedas),
        "comision": round(comision, 2),
        "neto": round(neto, 2),
        "moneda_neto": _una(monedas_neto),
        # Cobros sin detalle de liquidacion: si es > 0, el neto esta incompleto
        # y hay que apretar Reconciliar para completarlo.
        "sin_liquidacion": sin_datos,
    }

    # --- Lo que Stripe YA deposito en el banco de la LLC ---------------------
    # Esto es lo mas fuerte para reclamar: no es lo que decimos nosotros que se
    # cobro, es lo que Stripe informa que giro, con fecha.
    saldo, giros, error_stripe, cuenta = None, [], None, None
    if settings.STRIPE_SECRET_KEY:
        try:
            import stripe
            stripe.api_key = settings.STRIPE_SECRET_KEY
            from checkout import _sd
            b = _sd(stripe.Balance.retrieve())

            def _plata(lista):
                return [{"monto": (_sd(x).get("amount") or 0) / 100,
                         "moneda": (_sd(x).get("currency") or "").upper()}
                        for x in (lista or [])]

            saldo = {"disponible": _plata(b.get("available")),
                     "pendiente": _plata(b.get("pending"))}
            # --- Por que NO sale la plata --------------------------------
            # Un saldo disponible sin giros puede ser tres cosas muy distintas,
            # y el reclamo cambia segun cual sea: que no haya banco cargado,
            # que los giros esten en manual (alguien tiene que apretar), o que
            # Stripe tenga la cuenta trabada pidiendo documentacion.
            try:
                acc = _sd(stripe.Account.retrieve())
                sched = _sd(_sd(_sd(acc.get("settings")).get("payouts")).get("schedule"))
                reqs = _sd(acc.get("requirements"))
                bancos = []
                for ext in (_sd(acc.get("external_accounts")).get("data") or []):
                    ext = _sd(ext)
                    bancos.append({
                        "banco": ext.get("bank_name") or ext.get("brand"),
                        "ultimos4": ext.get("last4"),
                        "pais": ext.get("country"),
                        "moneda": (ext.get("currency") or "").upper(),
                    })
                cuenta = {
                    "pais": acc.get("country"),
                    "cobra": bool(acc.get("charges_enabled")),
                    "puede_girar": bool(acc.get("payouts_enabled")),
                    # "manual" = la plata se queda hasta que alguien la saque.
                    "frecuencia": sched.get("interval"),
                    "demora_dias": sched.get("delay_days"),
                    "bancos": bancos,
                    # Lo que Stripe esta esperando para destrabar la cuenta.
                    "pendiente": (reqs.get("currently_due") or [])[:8],
                    "vence": reqs.get("current_deadline"),
                    "motivo_freno": reqs.get("disabled_reason"),
                }
            except Exception:  # noqa: BLE001 - dato extra, no puede tumbar la vista
                log.exception("dinero: no se pudo leer la cuenta de Stripe")
                cuenta = None

            for po in (_sd(stripe.Payout.list(limit=25)).get("data") or []):
                po = _sd(po)
                llega = po.get("arrival_date")
                giros.append({
                    "id": po.get("id"),
                    "monto": (po.get("amount") or 0) / 100,
                    "moneda": (po.get("currency") or "").upper(),
                    "estado": po.get("status"),
                    "llega_el": (datetime.fromtimestamp(int(llega), tz=timezone.utc)
                                 .isoformat() if llega else None),
                })
        except Exception as exc:  # noqa: BLE001 - la pantalla igual tiene que abrir
            log.exception("dinero: no se pudo consultar Stripe")
            error_stripe = f"{type(exc).__name__}: {str(exc)[:140]}"

    return {"resumen": resumen, "cobros": cobros, "saldo_stripe": saldo,
            "giros_al_banco": giros, "cuenta_stripe": cuenta,
            "error_stripe": error_stripe, "desde": desde, "hasta": hasta}


@router.get("/dinero/export")
def dinero_export(desde: Optional[str] = None, hasta: Optional[str] = None,
                  db: Session = Depends(get_db)):
    """El mismo detalle en CSV, para mandarselo a la LLC como reclamo."""
    import csv
    datos = dinero(desde=desde, hasta=hasta, db=db)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";")
    w.writerow(["Pedido", "Fecha", "Cliente", "Email", "Bruto", "Moneda",
                "Comision Stripe", "Neto", "Moneda neto", "Disponible el",
                "Medio", "ID de cobro en Stripe", "Recibo oficial"])
    for c in datos["cobros"]:
        w.writerow([c["pedido"], (c["fecha"] or "")[:19].replace("T", " "),
                    c["cliente"] or "", c["email"] or "", c["bruto"],
                    c["moneda"] or "",
                    c["comision"] if c["comision"] is not None else "",
                    c["neto"] if c["neto"] is not None else "",
                    c["moneda_neto"] or "", (c["disponible_el"] or "")[:10],
                    c["medio"] or "", c["cobro_id"] or "", c["recibo_url"] or ""])
    r = datos["resumen"]
    w.writerow([])
    w.writerow(["TOTALES", str(r["cantidad"]) + " cobros", "", "",
                r["bruto"], r["moneda"] or "", r["comision"], r["neto"],
                r["moneda_neto"] or "", "", "", "", ""])
    nombre = "cobros_miamiimport_" + (desde or "inicio") + "_" + (hasta or "hoy") + ".csv"
    # BOM para que Excel abra bien los acentos.
    return StreamingResponse(
        iter(["\ufeff" + buf.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="' + nombre + '"'})


@router.post("/orders/{oid}/refund")
def refund_order(oid: int, db: Session = Depends(get_db)):
    """Reembolsa el pago de Stripe asociado a la orden y marca el pedido."""
    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    # Validar el ESTADO antes que la config: reembolsar algo no cobrado es un
    # error de operación, no de configuración, y el mensaje tiene que decirlo.
    # Sin este control se llamaba a Stripe sobre órdenes pendientes o ya
    # reembolsadas y el error crudo de Stripe volvía al cliente.
    if o.payment_status == "refunded":
        return {"ok": True, "already": True, "status": "refunded"}
    if o.payment_status != "paid":
        raise HTTPException(409, f"La orden no está pagada (estado: {o.payment_status})")
    if not settings.STRIPE_SECRET_KEY:
        raise HTTPException(503, "Stripe no está configurado")

    from core.models import AuditLog, Payment
    payment = (
        db.query(Payment).filter(Payment.order_id == o.id,
                                 Payment.stripe_payment_intent_id.isnot(None))
        .order_by(Payment.id.desc()).first()
    )
    if not payment or not payment.stripe_payment_intent_id:
        raise HTTPException(400, "No hay pago de Stripe para reembolsar")
    import stripe
    stripe.api_key = settings.STRIPE_SECRET_KEY
    try:
        refund = stripe.Refund.create(
            payment_intent=payment.stripe_payment_intent_id,
            # Doble click = un solo reembolso, no dos.
            idempotency_key=f"refund-order-{o.id}",
        )
    except stripe.StripeError as e:
        log.exception("Stripe rechazó el reembolso de la orden %s", o.id)
        raise HTTPException(502, "No se pudo reembolsar. Revisá el panel de Stripe.") from e

    payment.status = "refunded"
    o.payment_status = "refunded"
    o.status = "refunded"
    # Reponer stock una sola vez. Dejar stock_reserved en True hacía que un
    # `canceled` tardío volviera a sumar las mismas unidades (stock fantasma).
    if o.stock_reserved:
        for it in o.items:
            v = db.get(Variant, it.variant_id, with_for_update=True) if it.variant_id else None
            if v:
                v.stock = (v.stock or 0) + it.quantity
        o.stock_reserved = False
    db.add(AuditLog(action="order_refunded", entity="order", entity_id=str(o.id)))
    db.commit()
    return {"ok": True, "refund_id": refund.id, "status": "refunded"}


@router.get("/orders")
def list_orders(per_page: int = 50, page: int = 1, status: Optional[str] = None,
                db: Session = Depends(get_db)):
    # El barrido de reservas abandonadas corría SOLO al arrancar un checkout
    # nuevo. Si no entraba otra venta, las prendas de un carrito abandonado
    # quedaban descontadas por días y figuraban agotadas. Abrir Pedidos es el
    # otro momento natural para limpiarlas, y es lo que Diego hace todo el día.
    try:
        from checkout import _reap_abandoned_reservations
        _reap_abandoned_reservations(db)
    except Exception:  # noqa: BLE001 — la lista tiene que abrir igual
        log.exception("No se pudo barrer las reservas abandonadas")
        db.rollback()

    # selectinload: el serializer recorre `o.items` de cada pedido. Sin esto
    # SQLAlchemy los pedia de a UNO — 22 pedidos = 22 viajes de ida y vuelta a
    # Sao Paulo (~180 ms cada uno) y la pantalla de Pedidos tardaba 4,7 s.
    query = db.query(Order).options(selectinload(Order.items),
                                    selectinload(Order.payments))
    if status:
        query = query.filter(Order.payment_status == status)
    items = query.order_by(Order.created_at.desc()).offset((page - 1) * per_page).limit(per_page).all()
    return [order_to_tn(o) for o in items]


@router.get("/stats")
def stats(db: Session = Depends(get_db)):
    # Para los numeros del tablero alcanza con las variantes (de ahi sale el
    # stock) y el nombre. Traer ademas las imagenes y las categorias de los 261
    # productos era arrastrar miles de filas al pedo en cada apertura.
    productos = db.query(Product).options(selectinload(Product.variants)).all()
    total_productos = len(productos)
    total_publicados = sum(1 for p in productos if p.published)
    total_variantes = db.query(Variant).count()
    total_stock = db.query(func.coalesce(func.sum(Variant.stock), 0)).scalar() or 0
    productos_sin_stock = sum(1 for p in productos if p.total_stock == 0)

    pedidos = db.query(Order).options(selectinload(Order.items)).all()
    total_pedidos = len(pedidos)
    total_facturado = float(sum((o.total or 0) for o in pedidos))
    pedidos_pagados = sum(1 for o in pedidos if o.payment_status == "paid")
    pedidos_pendientes = sum(1 for o in pedidos if o.payment_status == "pending")

    contar = {}
    for o in pedidos:
        for it in o.items:
            if it.product_id:
                contar[it.product_id] = contar.get(it.product_id, 0) + it.quantity
    top = sorted(contar.items(), key=lambda x: -x[1])[:10]
    name_by_id = {p.id: p.name for p in productos}
    top_named = [{"product_id": pid, "name": name_by_id.get(pid, "?"), "vendidos": q} for pid, q in top]

    stock_bajo = [
        {"id": p.id, "name": p.name, "brand": p.brand, "stock": p.total_stock}
        for p in productos if p.total_stock <= 1
    ][:20]

    return {
        "productos": {
            "total": total_productos, "publicados": total_publicados,
            "sin_stock": productos_sin_stock, "variantes": total_variantes,
            "stock_total": int(total_stock),
        },
        "pedidos": {
            "total": total_pedidos, "pagados": pedidos_pagados, "pendientes": pedidos_pendientes,
            "facturado_total": total_facturado,
            "ticket_promedio": (total_facturado / total_pedidos) if total_pedidos else 0,
        },
        "top_vendidos": top_named,
        "stock_bajo": stock_bajo,
    }


# --------------------------------------------------------------------------- #
# Precios USD (ahora 100% sobre la DB propia)
# --------------------------------------------------------------------------- #
@router.get("/usd_prices")
def usd_prices_get(db: Session = Depends(get_db)):
    prices = {}
    for p in _productos_completos(db).all():
        v = p.variants[0] if p.variants else None
        if v and v.usd_price is not None:
            prices[str(p.id)] = float(v.usd_price)
    return {"prices": prices, "rate": current_usd_rate(db)}


@router.post("/usd_prices")
def usd_prices_save(body: dict, db: Session = Depends(get_db)):
    # La cotización se guarda ACÁ, en la tabla settings, que es de donde la lee
    # current_usd_rate(). El botón "Guardar cotización" la mandaba a
    # /api/bot_config (un JSON suelto del bot) y el recálculo nunca la veía:
    # guardabas 1500 y seguía multiplicando por 1410.
    if body.get("rate") is not None:
        try:
            rate = Decimal(str(body["rate"]))
        except (InvalidOperation, TypeError, ValueError):
            raise HTTPException(400, "Cotización inválida")
        if not rate.is_finite() or rate <= 0 or rate > Decimal("1000000"):
            raise HTTPException(400, "Cotización fuera de rango")
        s = db.get(Setting, "usd_rate")
        if not s:
            s = Setting(key="usd_rate")
            db.add(s)
        s.value = {"rate": float(rate)}

    prices = body.get("prices") or {}
    saved = 0
    for pid, usd in prices.items():
        try:
            usd_d = Decimal(str(usd))
        except Exception:
            continue
        p = db.get(Product, int(pid))
        if not p:
            continue
        for v in p.variants:
            v.usd_price = usd_d
        saved += 1
    db.commit()
    return {"ok": True, "saved_count": saved, "rate": current_usd_rate(db)}


@router.post("/usd_prices/from_current")
def usd_prices_seed(db: Session = Depends(get_db)):
    rate = Decimal(str(current_usd_rate(db)))
    if rate <= 0:
        raise HTTPException(400, "USD rate inválido")
    count = 0
    for p in _productos_completos(db).all():
        for v in p.variants:
            if v.price:
                v.usd_price = (v.price / rate).quantize(Decimal("0.01"))
        count += 1
    db.commit()
    return {"ok": True, "count": count, "rate": float(rate)}


@router.post("/usd_prices/sync_to_tiendanube")
def usd_prices_sync(db: Session = Depends(get_db)):
    """Recalcula ARS = USD × rate en TODAS las variantes (ya no toca Tiendanube,
    actualiza nuestra propia base — el nombre se mantiene por compatibilidad del frontend)."""
    rate = Decimal(str(current_usd_rate(db)))
    updated_products = updated_variants = 0
    for p in _productos_completos(db).all():
        touched = False
        for v in p.variants:
            if v.usd_price is not None:
                v.price = (v.usd_price * rate).quantize(Decimal("0.01"))
                updated_variants += 1
                touched = True
        if touched:
            updated_products += 1
    db.commit()
    return {"ok": True, "updated_products": updated_products,
            "updated_variants": updated_variants, "rate": float(rate)}


# --------------------------------------------------------------------------- #
# Export Excel (desde la DB)
# --------------------------------------------------------------------------- #
@router.get("/export/excel")
def export_excel(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    productos = _productos_completos(db).order_by(Product.id.desc()).all()
    pedidos = db.query(Order).order_by(Order.created_at.desc()).limit(1000).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "productos"
    ws.append(["id", "nombre", "marca", "publicado", "url", "stock_total",
               "variantes", "precio_min", "categorías"])
    for p in productos:
        cats = " | ".join(c.name for c in p.categories)
        ws.append([p.id, p.name, p.brand, p.published,
                   f"{settings.STORE_BASE_URL}/productos/{p.handle}/",
                   p.total_stock, len(p.variants),
                   float(p.min_price) if p.min_price else 0, cats])

    wsv = wb.create_sheet("variantes")
    wsv.append(["product_id", "product_name", "variant_id", "sku", "talle", "precio", "stock"])
    for p in productos:
        for v in p.variants:
            wsv.append([p.id, p.name, v.id, v.sku, v.value,
                        float(v.price) if v.price else 0, v.stock])

    wsp = wb.create_sheet("pedidos")
    wsp.append(["id", "numero", "fecha", "estado", "pago", "total", "moneda",
                "cliente", "email", "telefono", "productos"])
    for o in pedidos:
        prods = " | ".join(f"{it.product_name} x{it.quantity}" for it in o.items)
        wsp.append([o.id, o.number, o.created_at.isoformat() if o.created_at else "",
                    o.status, o.payment_status, float(o.total or 0), o.currency,
                    o.contact_name, o.email, o.contact_phone, prods])

    for sh in (ws, wsv, wsp):
        for cell in sh[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="111111", end_color="111111", fill_type="solid")
        sh.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=miami_import_export_{stamp}.xlsx"},
    )


# --------------------------------------------------------------------------- #
# Reservas de productos "a pedido"
# --------------------------------------------------------------------------- #
@router.get("/reservas")
def list_reservas(db: Session = Depends(get_db)):
    """Todas las reservas, las pendientes primero y dentro de cada grupo las
    más nuevas arriba."""
    orden = case((Reserva.status == "pendiente", 0),
                 (Reserva.status == "avisado", 1),
                 (Reserva.status == "entregado", 2),
                 else_=3)
    reservas = (db.query(Reserva)
                .order_by(orden, Reserva.created_at.desc())
                .all())
    return [reserva_to_dict(r) for r in reservas]


@router.post("/reservas")
def create_reserva(body: dict, db: Session = Depends(get_db)):
    """Registra que un cliente quiere un producto (y opcionalmente un talle).

    El producto tiene que ser uno "a pedido". Se guardan copias del nombre y el
    talle para que la reserva se lea aunque el producto se borre después.
    """
    nombre = str(body.get("customer_name") or "").strip()[:255]
    if not nombre:
        raise HTTPException(400, "Falta el nombre del cliente")

    pid = body.get("product_id")
    prod = db.get(Product, int(pid)) if pid not in (None, "") else None
    if not prod:
        raise HTTPException(404, "Producto no encontrado")
    if not prod.a_pedido:
        raise HTTPException(400, "Solo se reservan productos de la pestaña 'A pedido'")

    talle = None
    variant_id = body.get("variant_id")
    if variant_id not in (None, ""):
        v = db.get(Variant, int(variant_id))
        if not v or v.product_id != prod.id:
            raise HTTPException(400, "El talle no es de este producto")
        variant_id = v.id
        talle = v.value
    else:
        variant_id = None

    r = Reserva(
        product_id=prod.id,
        variant_id=variant_id,
        product_name=prod.name,
        talle=talle,
        customer_name=nombre,
        customer_phone=(str(body.get("customer_phone") or "").strip()[:50] or None),
        notes=(str(body.get("notes") or "").strip()[:1000] or None),
        status="pendiente",
    )
    db.add(r)
    db.commit()
    db.refresh(r)
    return reserva_to_dict(r)


@router.post("/reservas/{rid}/status")
def set_reserva_status(rid: int, body: dict, db: Session = Depends(get_db)):
    r = db.get(Reserva, rid)
    if not r:
        raise HTTPException(404, "Reserva no encontrada")
    nuevo = str(body.get("status") or "").strip().lower()
    if nuevo not in RESERVA_STATUSES:
        raise HTTPException(400, f"Estado inválido: {nuevo or '(vacío)'}")
    r.status = nuevo
    db.commit()
    return reserva_to_dict(r)


@router.delete("/reservas/{rid}")
def delete_reserva(rid: int, db: Session = Depends(get_db)):
    r = db.get(Reserva, rid)
    if not r:
        raise HTTPException(404, "Reserva no encontrada")
    db.delete(r)
    db.commit()
    return {"ok": True}


# --------------------------------------------------------------------------- #
# LA WEB — la home editable por Diego (hero, vitrina, marcas, valores, cierre)
# --------------------------------------------------------------------------- #
@router.get("/web/home")
def web_home_get(db: Session = Depends(get_db)):
    """Config actual + si hay motor de recorte disponible (el panel se adapta)."""
    return {"config": get_home_config(db), "recorte_ia": recorte_disponible()}


@router.put("/web/home")
def web_home_put(body: dict, db: Session = Depends(get_db)):
    """Guarda uno o varios bloques. Se mergea: lo que no venga, no se toca."""
    if not isinstance(body, dict) or not body:
        raise HTTPException(400, "No llegó ningún cambio")
    bloques = {k: v for k, v in body.items() if k in HOME_DEFAULTS}
    if not bloques:
        raise HTTPException(400, "Ningún bloque válido en el pedido")
    return {"config": save_home_config(db, bloques)}


@router.post("/web/home/reset")
def web_home_reset(db: Session = Depends(get_db)):
    """Vuelve la home a como vino de fábrica."""
    return {"config": reset_home_config(db)}


@router.post("/web/imagen")
async def web_imagen(
    file: UploadFile = File(...),
    quitar_fondo: bool = Form(True),
    destino: str = Form("vitrina"),
):
    """Sube una foto para la home y (si se pide) le saca el fondo con IA.

    Si el recorte falla, la foto se guarda igual TAL CUAL y se avisa en la
    respuesta: es preferible que Diego vea su imagen con fondo y decida, a que
    la subida se caiga y no pueda cargar nada.
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(400, "Archivo vacío")
    if len(raw) > MAX_IMAGE_BYTES:
        raise HTTPException(413, "La imagen es demasiado grande (máx. 8 MB)")

    destino = re.sub(r"[^a-z0-9_-]", "", (destino or "vitrina").lower()) or "vitrina"
    data, motor, aviso = raw, None, None

    if quitar_fondo:
        if not recorte_disponible():
            aviso = ("No hay IA de recorte configurada (falta GEMINI_API_KEY o "
                     "FAL_KEY). La foto se subió con su fondo original.")
        else:
            try:
                data, motor = quitar_fondo_ia(raw)
            except Exception as exc:  # noqa: BLE001
                log.warning("recorte falló: %s", exc)
                aviso = ("No se pudo recortar el fondo automáticamente; la foto "
                         "quedó con su fondo. Probá con otra o subí un PNG ya recortado.")
    if motor is None and not quitar_fondo:
        data = comprimir_imagen(raw)

    ext = "png" if motor else (Path(file.filename or "").suffix.lower().lstrip(".") or "jpg")
    if ext not in ("png", "jpg", "jpeg", "webp"):
        ext = "jpg"
    rel = f"home/{destino}/{secrets.token_hex(12)}.{ext}"
    tipo = "image/png" if ext == "png" else f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"

    if storage.is_enabled():
        try:
            url = storage.upload_bytes(data, rel, tipo)
        except RuntimeError as exc:
            raise HTTPException(502, f"No se pudo guardar la imagen: {exc}")
    else:
        dest = STORE_STATIC / rel
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(data)
        url = f"/static/{rel}"

    return {"ok": True, "url": url, "motor": motor, "aviso": aviso}


# --------------------------------------------------------------------------- #
# ETIQUETA DE DESPACHO — hoja imprimible con los datos del envío
# --------------------------------------------------------------------------- #
@router.get("/orders/{oid}/etiqueta")
def etiqueta_pedido(oid: int, db: Session = Depends(get_db)):
    """Etiqueta lista para imprimir y pegar en el paquete (Ctrl+P desde el panel).

    Diego despachaba copiando los datos a mano. Esto arma la hoja con
    remitente, destinatario, teléfono, CP y el detalle con TALLE.
    """
    from fastapi.responses import HTMLResponse as _HTML

    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    d = o.shipping_address or {}
    tel = o.contact_phone or d.get("phone") or ""
    piso = f", {d.get('floor')}" if d.get("floor") else ""
    filas = "".join(
        f"<tr><td>{html.escape(it.product_name)}</td>"
        f"<td class='talle'>{html.escape(it.variant_value or '—')}</td>"
        f"<td>×{it.quantity}</td></tr>"
        for it in o.items
    )
    pagina = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"/>
<title>Etiqueta pedido #{o.number}</title>
<style>
  @page {{ size: A5 landscape; margin: 8mm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; color: #111; margin: 0; padding: 16px; }}
  .caja {{ border: 2.5px solid #111; border-radius: 10px; padding: 18px 22px; max-width: 640px; }}
  .fila {{ display: flex; justify-content: space-between; align-items: baseline;
           border-bottom: 1.5px solid #111; padding-bottom: 10px; margin-bottom: 12px; }}
  .marca {{ font-weight: 800; letter-spacing: .28em; font-size: 15px; }}
  .nro {{ font-size: 22px; font-weight: 800; }}
  .rotulo {{ font-size: 9px; letter-spacing: .18em; text-transform: uppercase; color: #666; margin: 10px 0 2px; }}
  .dest {{ font-size: 19px; font-weight: 800; }}
  .dir {{ font-size: 15px; line-height: 1.5; }}
  .cp {{ font-size: 26px; font-weight: 800; border: 2px solid #111; border-radius: 8px;
         display: inline-block; padding: 2px 14px; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 6px; font-size: 13px; }}
  td {{ padding: 4px 6px 4px 0; border-bottom: 1px dashed #bbb; }}
  .talle {{ font-weight: 800; font-size: 15px; white-space: nowrap; }}
  .pie {{ margin-top: 12px; font-size: 11px; color: #555; }}
  .noprint {{ margin: 14px 0 0; }}
  @media print {{ .noprint {{ display: none; }} }}
</style></head><body>
<div class="caja">
  <div class="fila"><span class="marca">MIAMI IMPORT</span><span class="nro">PEDIDO #{o.number}</span></div>
  <div class="rotulo">Destinatario</div>
  <div class="dest">{html.escape(o.contact_name or 'Sin nombre')}</div>
  <div class="dir">{html.escape((d.get('street') or '') + ' ' + (d.get('number') or ''))}{html.escape(piso)}<br/>
    {html.escape(d.get('city') or '')}{', ' + html.escape(d.get('province') or '') if d.get('province') else ''}
    &nbsp;·&nbsp; Tel: {html.escape(tel or 's/d')}</div>
  {f'<div class="cp">CP {html.escape(str(d.get("zipcode")))}</div>' if d.get('zipcode') else ''}
  <div class="rotulo">Contenido</div>
  <table>{filas}</table>
  <div class="pie">Remitente: MIAMI IMPORT · miamiimport.com.ar · WhatsApp 11 6232-1391</div>
</div>
<button class="noprint" onclick="window.print()">🖨️ Imprimir</button>
</body></html>"""
    return _HTML(pagina)


@router.get("/orders/{oid}/comprobante")
def comprobante_pedido(oid: int, db: Session = Depends(get_db)):
    """Comprobante de la venta, para compartir y archivar.

    Diego (audio 19-ago): *"necesito los comprobantes de cada compra para ir
    compartiéndolos y tener un respaldo de las transacciones"*. Incluye el ID
    de la transacción en Stripe, que es lo que permite rastrear la plata.
    """
    from fastapi.responses import HTMLResponse as _HTML
    from core.models import Payment

    o = db.get(Order, oid)
    if not o:
        raise HTTPException(404, "Pedido no encontrado")
    pago = (db.query(Payment).filter(Payment.order_id == o.id)
            .order_by(Payment.id.desc()).first())
    d = o.shipping_address or {}
    pagado = o.payment_status == "paid"

    def _money(x) -> str:
        return "$ " + f"{float(x or 0):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")

    filas = "".join(
        f"<tr><td>{html.escape(it.product_name)}"
        f"{f'<br><small>Talle {html.escape(it.variant_value)}</small>' if it.variant_value else ''}</td>"
        f"<td class='c'>{it.quantity}</td>"
        f"<td class='r'>{_money(it.unit_price)}</td>"
        f"<td class='r'>{_money((it.unit_price or 0) * it.quantity)}</td></tr>"
        for it in o.items
    )
    fecha = o.created_at.strftime("%d/%m/%Y %H:%M") if o.created_at else "—"
    tel = o.contact_phone or d.get("phone") or "—"
    envio = ("Retiro / venta en el local" if d.get("canal") == "local" else
             " · ".join(p for p in [f"{d.get('street','')} {d.get('number','')}".strip(),
                                    d.get("floor") or "", d.get("city") or "",
                                    d.get("province") or "",
                                    f"CP {d.get('zipcode')}" if d.get("zipcode") else ""] if p)
             or "—")
    sello = ("PAGADO" if pagado else "PAGO PENDIENTE")
    color = "#1a7f37" if pagado else "#b45309"
    trx = (pago.stripe_payment_intent_id if pago and pago.stripe_payment_intent_id
           else ("venta de mostrador" if d.get("canal") == "local" else "—"))

    # Rastro verificable del cobro: sin esto el comprobante dice "PAGADO"
    # porque lo decimos nosotros. Con el id del cobro y el recibo de Stripe,
    # cualquiera puede confirmarlo por afuera.
    partes = []
    if pago and getattr(pago, "stripe_charge_id", None):
        partes.append(f"<b>Cobro Stripe:</b> {html.escape(pago.stripe_charge_id)}")
    if pago and getattr(pago, "card_brand", None) and getattr(pago, "card_last4", None):
        partes.append(f"<b>Tarjeta:</b> {html.escape(pago.card_brand.upper())} ....{html.escape(pago.card_last4)}")
    elif pago and getattr(pago, "metodo", None):
        from panel.serializers import _MEDIOS
        medio = _MEDIOS.get(pago.metodo, pago.metodo.replace("_", " ").title())
        partes.append(f"<b>Medio de pago:</b> {html.escape(medio)}")
    if pago and getattr(pago, "paid_at", None):
        partes.append(f"<b>Acreditado:</b> {pago.paid_at.strftime('%d/%m/%Y %H:%M')} hs")
    if pago and getattr(pago, "receipt_url", None):
        partes.append(f"<b>Recibo oficial:</b> {html.escape(pago.receipt_url)}")
    rastro = ("<br/>".join(partes) + "<br/>") if partes else ""

    pagina = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"/>
<title>Comprobante #{o.number} — MIAMI IMPORT</title>
<style>
  @page {{ size: A4; margin: 14mm; }}
  body {{ font-family: Arial, Helvetica, sans-serif; color: #111; margin: 0; padding: 24px; }}
  .hoja {{ max-width: 720px; margin: 0 auto; }}
  .top {{ display: flex; justify-content: space-between; align-items: flex-start;
          border-bottom: 3px solid #111; padding-bottom: 14px; }}
  .marca {{ font-weight: 800; letter-spacing: .3em; font-size: 17px; }}
  .sub {{ font-size: 11px; color: #666; margin-top: 4px; }}
  .nro {{ text-align: right; }}
  .nro b {{ font-size: 22px; }}
  .sello {{ display: inline-block; margin-top: 6px; padding: 5px 14px; border-radius: 6px;
            color: #fff; font-weight: 800; font-size: 12px; letter-spacing: .1em;
            background: {color}; }}
  .bloques {{ display: flex; gap: 28px; margin: 20px 0 6px; }}
  .bloque {{ flex: 1; }}
  .rot {{ font-size: 9px; letter-spacing: .18em; text-transform: uppercase;
          color: #888; margin-bottom: 4px; }}
  .dato {{ font-size: 13px; line-height: 1.6; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 16px; font-size: 13px; }}
  th {{ text-align: left; font-size: 9px; letter-spacing: .16em; text-transform: uppercase;
        color: #888; border-bottom: 1.5px solid #111; padding: 6px 4px; }}
  td {{ padding: 9px 4px; border-bottom: 1px solid #e5e5e5; vertical-align: top; }}
  td small {{ color: #777; }}
  .c {{ text-align: center; }} .r {{ text-align: right; white-space: nowrap; }}
  .total {{ text-align: right; font-size: 19px; font-weight: 800; margin-top: 14px; }}
  .trx {{ margin-top: 20px; padding: 10px 12px; background: #f6f6f6; border-radius: 8px;
          font-size: 11px; color: #555; word-break: break-all; }}
  .pie {{ margin-top: 24px; padding-top: 12px; border-top: 1px solid #ddd;
          font-size: 11px; color: #777; text-align: center; }}
  .noprint {{ margin-top: 18px; text-align: center; }}
  @media print {{ .noprint {{ display: none; }} }}
</style></head><body><div class="hoja">
  <div class="top">
    <div><div class="marca">MIAMI IMPORT</div>
      <div class="sub">Indumentaria original importada<br/>miamiimport.com.ar · WhatsApp 11 6232-1391</div></div>
    <div class="nro"><div class="rot">Comprobante</div><b>#{o.number}</b>
      <div class="sub">{fecha}</div><div class="sello">{sello}</div></div>
  </div>

  <div class="bloques">
    <div class="bloque"><div class="rot">Cliente</div>
      <div class="dato"><b>{html.escape(o.contact_name or '—')}</b><br/>
        {html.escape(o.email or '—')}<br/>{html.escape(tel)}</div></div>
    <div class="bloque"><div class="rot">Envío</div>
      <div class="dato">{html.escape(envio)}</div></div>
  </div>

  <table>
    <tr><th>Producto</th><th class="c">Cant.</th><th class="r">Precio</th><th class="r">Subtotal</th></tr>
    {filas}
  </table>
  <div class="total">TOTAL &nbsp; {_money(o.total)} {html.escape(o.currency or 'ARS')}</div>

  <div class="trx"><b>Transacción:</b> {html.escape(str(trx))}<br/>
    {rastro}<b>Estado del pago:</b> {sello}{'' if pagado else ' — la plata todavía no ingresó'}</div>

  <div class="pie">Comprobante generado por el sistema de MIAMI IMPORT · {fecha}</div>
  <div class="noprint"><button onclick="window.print()">🖨️ Imprimir / Guardar PDF</button></div>
</div></body></html>"""
    return _HTML(pagina)


# --------------------------------------------------------------------------- #
# MANTENIMIENTO: ordenar el catálogo heredado de Tiendanube
# --------------------------------------------------------------------------- #
@router.post("/catalogo/ordenar")
def catalogo_ordenar(body: dict | None = None, db: Session = Depends(get_db)):
    """Normaliza marcas y categoriza los productos que quedaron sueltos.

    La migración desde Tiendanube (y las cargas nuevas) dejaron productos sin
    categoría — invisibles en el menú de la tienda — y la misma marca escrita
    de varias formas, lo que partía su catálogo en pedazos.

    Con `{"dry": true}` NO toca nada: solo informa qué haría. Es lo primero
    que hay que correr.
    """
    dry = bool((body or {}).get("dry"))
    forzar = bool((body or {}).get("forzar"))

    productos = db.query(Product).order_by(Product.id).all()
    marcas_cambiadas, categorizados, sin_resolver = [], [], []

    for prod in productos:
        antes_marca, antes_cats = prod.brand, len(prod.categories or [])
        try:
            cambios = categorizar_producto(db, prod, forzar=forzar)
        except Exception as exc:  # noqa: BLE001
            log.exception("ordenar: falló el producto %s", prod.id)
            sin_resolver.append({"id": prod.id, "nombre": prod.name,
                                 "motivo": f"error: {str(exc)[:120]}"})
            continue

        if "marca" in cambios:
            marcas_cambiadas.append({"id": prod.id, "nombre": prod.name,
                                     "cambio": cambios["marca"]})
        if "categoria" in cambios:
            categorizados.append({"id": prod.id, "nombre": prod.name,
                                  "categoria": cambios["categoria"]})
        if "sin_categoria" in cambios and not antes_cats:
            sin_resolver.append({"id": prod.id, "nombre": prod.name,
                                 "motivo": cambios["sin_categoria"]})
        if dry:
            # deshacer lo que la función haya tocado en memoria
            prod.brand = antes_marca

    if dry:
        db.rollback()
    else:
        try:
            db.commit()
        except Exception as exc:  # noqa: BLE001
            db.rollback()
            log.exception("ordenar: no se pudo guardar")
            raise HTTPException(500, f"No se pudo guardar: {exc}") from exc

    return {
        "ok": True,
        "dry": dry,
        "revisados": len(productos),
        "marcas_normalizadas": marcas_cambiadas,
        "categorizados": categorizados,
        "sin_resolver": sin_resolver,
        "resumen": (f"{len(marcas_cambiadas)} marcas normalizadas · "
                    f"{len(categorizados)} productos categorizados · "
                    f"{len(sin_resolver)} sin resolver"),
    }
